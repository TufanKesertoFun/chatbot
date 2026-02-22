# Doc_Intelligence.py
from fastapi import Depends, HTTPException
from config import get_settings, Settings
import tempfile
import os
import traceback
import pandas as pd
import io
import datetime
import uuid
import re
from typing import Optional, List, Dict, Any

# Import Document Intelligence dependencies
from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.core.credentials import AzureKeyCredential
from langchain.text_splitter import MarkdownHeaderTextSplitter
from langchain.schema import Document

# Import Blob Storage dependencies
from Managers.Storage_1.Storage import AsyncStorage
from logging_config import get_logger

# Import Prometheus metrics
from Managers.prometheus_metrics import track_document_processing, MemoryMonitor

# Import profiling utilities
from Managers.Document_Intelligence_3.profiling import profile_endpoint, Profiler

# Module logger
logger = get_logger(__name__)


class DocumentIntelligence:
    def __init__(self, settings: Settings = Depends(get_settings)):
        logger.debug("Initializing DocumentIntelligence manager")

        self.api_key = settings.document_intelligence_key
        self.endpoint = settings.document_intelligence_endpoint

        # Initialize the Document Intelligence client
        self.client = DocumentIntelligenceClient(
            endpoint=self.endpoint,
            credential=AzureKeyCredential(self.api_key)
        )

        # Initialize the Storage manager
        self.storage_manager = AsyncStorage(settings)

        logger.info("DocumentIntelligence manager initialized successfully")
    
    @track_document_processing("DocumentIntelligence")
    @profile_endpoint(name='analyze_and_split_document_from_blob', document_type='general')
    async def analyze_and_split_document_from_blob(self, container_name, blob_path, model="prebuilt-layout",
                                               headers_to_split_on=None, bot_id=None, user_id=None):
        """
        Analyze and split a document from blob storage using Document Intelligence for supported formats
        and custom processing for CSV/XLSX files

        Args:
            container_name: Azure Storage container name
            blob_path: Path to the blob within the container
            model: Document intelligence model to use
            headers_to_split_on: List of headers to split document on
            bot_id: ID of the bot for multi-tenancy filtering
            user_id: ID of the user for multi-tenancy filtering
        """
        try:
            # Download the blob content
            blob_content = await self.storage_manager.download_file(container_name, blob_path)
            
            # Get document name from blob path
            document_name = blob_path.split('/')[-1]
            file_extension = document_name.lower().split('.')[-1] if '.' in document_name else ''
            
            # Handle CSV and XLSX files separately
            if file_extension in ['csv', 'xlsx', 'xls']:
                return await self._process_spreadsheet_file(
                    blob_content, document_name, file_extension, 
                    container_name, blob_path, bot_id, user_id, headers_to_split_on
                )
            
            # Handle Word documents with enhanced processing
            elif file_extension in ['docx', 'doc']:
                return await self._process_word_document(
                    blob_content, document_name, container_name, blob_path, 
                    model, headers_to_split_on, bot_id, user_id
                )
            
            # For other formats, use standard Document Intelligence
            return await self._process_with_document_intelligence(
                blob_content, document_name, container_name, blob_path, 
                model, headers_to_split_on, bot_id, user_id
            )
            
        except Exception as e:
            logger.debug(f"Document analysis and splitting failed: {str(e)}")
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Document analysis and splitting failed: {str(e)}")

    async def _process_word_document(self, blob_content, document_name, container_name,
                                   blob_path, model, headers_to_split_on, bot_id, user_id):
        """Process Word documents with enhanced structure detection"""
        temp_file_path = None

        # Calculate document size
        doc_size_kb = len(blob_content) / 1024 if blob_content else 0

        with Profiler(
            name='process_word_document',
            document_type='docx',
            document_size_kb=doc_size_kb,
            pages_processed=None  # Will be set later
        ) as profiler:
            try:
                # Create a temporary file
                with tempfile.NamedTemporaryFile(delete=False, suffix='.docx') as temp_file:
                    temp_file.write(blob_content)
                    temp_file_path = temp_file.name

                # Use Document Intelligence to analyze the Word document
                with open(temp_file_path, "rb") as f:
                    file_content = f.read()
                    poller = self.client.begin_analyze_document(
                        model_id=model,
                        body=file_content,
                        content_type="application/octet-stream"
                    )
                    result = poller.result()

                # Extract document structure information
                document_info = self._extract_word_document_info(result)

                # Update profiler with page count
                if profiler.result:
                    profiler.pages_processed = document_info.get('total_pages', 0)

                # Process based on document structure
                if headers_to_split_on or self._has_clear_structure(result):
                    return await self._process_structured_word_document(
                        result, document_name, container_name, blob_path,
                        headers_to_split_on, bot_id, user_id, document_info
                    )
                else:
                    return await self._process_word_document_by_pages(
                        result, document_name, container_name, blob_path,
                        bot_id, user_id, document_info
                    )

            except Exception as e:
                logger.debug(f"Word document processing failed: {str(e)}")
                traceback.print_exc()
                raise HTTPException(status_code=500, detail=f"Word document processing failed: {str(e)}")

            finally:
                # Clean up temp file
                if temp_file_path and os.path.exists(temp_file_path):
                    try:
                        os.unlink(temp_file_path)
                    except Exception as e:
                        logger.debug(f"Warning: Could not delete temp file {temp_file_path}: {str(e)}")

    def _extract_word_document_info(self, result) -> Dict[str, Any]:
        """Extract structural information from Word document analysis"""
        info = {
            "total_pages": len(result.pages) if result.pages else 0,
            "has_tables": bool(result.tables),
            "table_count": len(result.tables) if result.tables else 0,
            "has_paragraphs": bool(result.paragraphs),
            "paragraph_count": len(result.paragraphs) if result.paragraphs else 0,
            "styles": set(),
            "potential_headers": []
        }
        
        # Analyze paragraphs for potential headers and styles
        if result.paragraphs:
            for para in result.paragraphs:
                if hasattr(para, 'role') and para.role:
                    info["styles"].add(para.role)
                    
                # Check if paragraph might be a header (short text, often at start of line)
                content = para.content.strip()
                if (len(content) < 100 and 
                    not content.endswith('.') and 
                    not content.endswith(',') and
                    len(content.split()) <= 10):
                    info["potential_headers"].append(content)
        
        return info

    def _has_clear_structure(self, result) -> bool:
        """Determine if the document has clear structural elements"""
        if not result.paragraphs:
            return False
            
        # Check for role-based structure (headers, titles, etc.)
        structured_roles = {'title', 'sectionHeading', 'pageHeader', 'pageFooter'}
        for para in result.paragraphs:
            if hasattr(para, 'role') and para.role in structured_roles:
                return True
                
        # Check for potential header patterns
        header_patterns = [
            r'^\d+\.\s+',      # "1. Introduction"
            r'^[A-Z\s]+:',     # "CHAPTER ONE:"
            r'^[A-Z][a-z\s]+$' # "Introduction" (capitalized words only)
        ]
        
        potential_headers = 0
        for para in result.paragraphs:
            content = para.content.strip()
            if len(content) < 100:  # Short lines are more likely to be headers
                for pattern in header_patterns:
                    if re.match(pattern, content):
                        potential_headers += 1
                        break
        
        return potential_headers >= 2  # At least 2 potential headers suggest structure

    async def _process_structured_word_document(self, result, document_name, container_name, 
                                              blob_path, headers_to_split_on, bot_id, user_id, document_info):
        """Process Word document by logical sections"""
        documents = []
        
        # Create document overview
        overview_content = self._create_word_document_overview(result, document_name, document_info)
        overview_doc = self._create_document_chunk(
            overview_content, container_name, blob_path, bot_id, user_id,
            chunk_number=1, context="document overview", page_number=1,
            file_type="docx", summary="Document structure and overview"
        )
        documents.append(overview_doc)
        
        # If specific headers are provided, use them for splitting
        if headers_to_split_on:
            return await self._split_by_custom_headers(
                result, documents, container_name, blob_path, 
                headers_to_split_on, bot_id, user_id
            )
        
        # Otherwise, split by detected structure
        return await self._split_by_detected_structure(
            result, documents, container_name, blob_path, 
            bot_id, user_id, document_info
        )

    async def _process_word_document_by_pages(self, result, document_name, container_name, 
                                            blob_path, bot_id, user_id, document_info):
        """Process Word document page by page when no clear structure is detected"""
        documents = []
        
        # Create document overview
        overview_content = self._create_word_document_overview(result, document_name, document_info)
        overview_doc = self._create_document_chunk(
            overview_content, container_name, blob_path, bot_id, user_id,
            chunk_number=1, context="document overview", page_number=1,
            file_type="docx", summary="Document structure and overview"
        )
        documents.append(overview_doc)
        
        # Process each page
        for page_index, page in enumerate(result.pages):
            if not page.lines:
                continue
                
            page_content = self._extract_page_content_with_structure(page, result)
            
            page_doc = self._create_document_chunk(
                page_content, container_name, blob_path, bot_id, user_id,
                chunk_number=page_index + 2, context="document page", 
                page_number=page.page_number, file_type="docx",
                summary=f"Page {page.page_number} content"
            )
            documents.append(page_doc)
        
        return documents

    def _create_word_document_overview(self, result, document_name, document_info) -> str:
        """Create an overview of the Word document"""
        overview = f"Document: {document_name}\n"
        overview += f"File Type: Microsoft Word Document\n"
        overview += f"Total Pages: {document_info['total_pages']}\n"
        overview += f"Total Paragraphs: {document_info['paragraph_count']}\n"
        
        if document_info['has_tables']:
            overview += f"Tables Found: {document_info['table_count']}\n"
        
        if document_info['styles']:
            overview += f"Document Styles: {', '.join(document_info['styles'])}\n"
        
        # Add content preview
        if result.paragraphs:
            overview += "\nDocument Structure:\n"
            for i, para in enumerate(result.paragraphs[:10]):  # First 10 paragraphs
                content = para.content.strip()
                if content:
                    role = getattr(para, 'role', 'paragraph')
                    overview += f"- {role}: {content[:100]}{'...' if len(content) > 100 else ''}\n"
        
        # Add table information if present
        if result.tables:
            overview += f"\nTables Summary:\n"
            for i, table in enumerate(result.tables[:3]):  # First 3 tables
                row_count = table.row_count if hasattr(table, 'row_count') else len(table.cells) // table.column_count if hasattr(table, 'column_count') else 'unknown'
                col_count = table.column_count if hasattr(table, 'column_count') else 'unknown'
                overview += f"- Table {i+1}: {row_count} rows, {col_count} columns\n"
        
        return overview

    def _extract_page_content_with_structure(self, page, result) -> str:
        """Extract page content while preserving structure information"""
        content = f"Page {page.page_number}:\n\n"
        
        # Get page lines
        page_lines = [line.content for line in page.lines] if page.lines else []
        
        # Try to identify and mark structural elements
        structured_content = []
        for line in page_lines:
            line_content = line.strip()
            if not line_content:
                continue
                
            # Check if this line might be a header
            if self._is_likely_header(line_content):
                structured_content.append(f"HEADER: {line_content}")
            else:
                structured_content.append(line_content)
        
        content += '\n'.join(structured_content)
        
        # Add any tables on this page
        if result.tables:
            page_tables = []
            for table in result.tables:
                # Check if table is on this page (simplified check)
                if hasattr(table, 'bounding_regions'):
                    for region in table.bounding_regions:
                        if region.page_number == page.page_number:
                            page_tables.append(table)
                            break
            
            if page_tables:
                content += f"\n\nTables on this page: {len(page_tables)}\n"
                for i, table in enumerate(page_tables):
                    content += f"\nTable {i+1}:\n"
                    content += self._extract_table_content(table)
        
        return content

    def _is_likely_header(self, text: str) -> bool:
        """Determine if text is likely a header"""
        if len(text) > 150:  # Too long to be a header
            return False
            
        # Common header patterns
        header_patterns = [
            r'^\d+\.\s+',           # "1. Introduction"
            r'^[A-Z\s]+:',          # "CHAPTER ONE:"
            r'^[A-Z][a-z\s]+$',     # "Introduction"
            r'^\d+\.\d+\s+',        # "1.1 Subsection"
            r'^[A-Z]{2,}',          # "INTRODUCTION"
        ]
        
        for pattern in header_patterns:
            if re.match(pattern, text):
                return True
                
        # Check if it's all caps (common for headers)
        words = text.split()
        if len(words) <= 5 and text.isupper():
            return True
            
        return False

    def _extract_table_content(self, table) -> str:
        """Extract table content in a readable format"""
        if not table.cells:
            return "Empty table\n"
            
        try:
            # Group cells by row
            rows = {}
            for cell in table.cells:
                row_index = cell.row_index
                if row_index not in rows:
                    rows[row_index] = {}
                rows[row_index][cell.column_index] = cell.content
            
            # Format as text table
            table_text = ""
            for row_index in sorted(rows.keys()):
                row = rows[row_index]
                row_text = " | ".join([row.get(col_index, "") for col_index in sorted(row.keys())])
                table_text += row_text + "\n"
            
            return table_text
            
        except Exception as e:
            logger.debug(f"Warning: Could not extract table content: {str(e)}")
            return "Table content could not be extracted\n"

    async def _split_by_custom_headers(self, result, documents, container_name, blob_path, 
                                     headers_to_split_on, bot_id, user_id):
        """Split document using custom header specifications"""
        # Convert all content to markdown-style text
        full_content = ""
        if result.paragraphs:
            for para in result.paragraphs:
                content = para.content.strip()
                if content:
                    # Try to detect header level
                    if self._is_likely_header(content):
                        # Add markdown header formatting
                        if re.match(r'^\d+\.\s+', content):
                            full_content += f"# {content}\n"
                        elif re.match(r'^\d+\.\d+\s+', content):
                            full_content += f"## {content}\n"
                        else:
                            full_content += f"# {content}\n"
                    else:
                        full_content += f"{content}\n"
        
        # Use langchain splitter
        text_splitter = MarkdownHeaderTextSplitter(headers_to_split_on=headers_to_split_on)
        
        try:
            splits = text_splitter.split_text(full_content)
            
            for split_index, split in enumerate(splits):
                split_keywords = []
                try:
                    from Managers.Document_Intelligence_3.semantic_chunker import extract_keywords
                    split_keywords = extract_keywords(split.page_content)
                except Exception:
                    pass
                
                split_doc = self._create_document_chunk(
                    split.page_content, container_name, blob_path, bot_id, user_id,
                    chunk_number=len(documents) + 1, context="header section",
                    page_number=split_index + 1, file_type="docx",
                    summary=f"Section {split_index + 1}",
                    keywords=split_keywords
                )
                documents.append(split_doc)
                
        except Exception as e:
            logger.debug(f"Error in custom header splitting: {str(e)}")
            # Fallback to page-based processing
            return await self._process_word_document_by_pages(
                result, container_name.split('/')[-1], container_name, blob_path, 
                bot_id, user_id, {"total_pages": len(result.pages)}
            )
        
        return documents

    async def _split_by_detected_structure(self, result, documents, container_name, blob_path, 
                                         bot_id, user_id, document_info):
        """Split document by automatically detected structure"""
        current_section = []
        section_header = "Introduction"
        section_number = 1
        
        if not result.paragraphs:
            return documents
            
        for para in result.paragraphs:
            content = para.content.strip()
            if not content:
                continue
                
            # Check if this is a new section header
            if self._is_likely_header(content):
                # Save previous section if it has content
                if current_section:
                    section_content = '\n'.join(current_section)
                    section_doc = self._create_document_chunk(
                        section_content, container_name, blob_path, bot_id, user_id,
                        chunk_number=len(documents) + 1, context="document section",
                        page_number=section_number, file_type="docx",
                        summary=f"Section: {section_header}"
                    )
                    documents.append(section_doc)
                
                # Start new section
                section_header = content
                section_number += 1
                current_section = [f"Section: {content}\n"]
            else:
                current_section.append(content)
        
        # Add the last section
        if current_section:
            section_content = '\n'.join(current_section)
            section_doc = self._create_document_chunk(
                section_content, container_name, blob_path, bot_id, user_id,
                chunk_number=len(documents) + 1, context="document section",
                page_number=section_number, file_type="docx",
                summary=f"Section: {section_header}"
            )
            documents.append(section_doc)
        
        return documents

    def _create_document_chunk(self, content: str, container_name: str, blob_path: str, 
                             bot_id: str, user_id: str, chunk_number: int, 
                             context: str, page_number: int, file_type: str, 
                             summary: str, keywords: List[str] = None) -> Dict[str, Any]:
        """Create a standardized document chunk"""
        
        # Extract keywords if not provided
        if keywords is None:
            try:
                from Managers.Document_Intelligence_3.semantic_chunker import extract_keywords
                keywords = extract_keywords(content)
            except Exception:
                keywords = []
        
        metadata = {
            "source": f"{container_name}/{blob_path}",
            "chunk": chunk_number,
            "page_number": page_number,
            "extraction_time": datetime.datetime.now().isoformat(),
            "context": context,
            "page_summary": summary,
            "keywords": ", ".join(keywords) if keywords else "",
            "bot_id": bot_id if bot_id else f"default-bot-{container_name}",
            "file_type": file_type
        }
        
        if user_id is not None:
            metadata["user_id"] = user_id
        
        return {
            "id": f"doc_{uuid.uuid4().hex[:8]}",
            "content": content,
            "metadata": metadata
        }

    async def _process_spreadsheet_file(self, blob_content, document_name, file_extension, 
                                       container_name, blob_path, bot_id, user_id, headers_to_split_on):
        """Process CSV and XLSX files with support for multiple sheets and hidden sheet detection"""
        try:
            documents = []
            
            if file_extension == 'csv':
                # CSV files only have one "sheet"
                df = await self._read_csv_file(blob_content, document_name)
                if df is not None and not df.empty:
                    sheet_documents = await self._process_single_sheet(
                        df, document_name, "Sheet1", container_name, blob_path, 
                        bot_id, user_id, file_extension, 1, 1
                    )
                    documents.extend(sheet_documents)
                
            elif file_extension in ['xlsx', 'xls']:
                # Excel files can have multiple sheets
                documents = await self._process_excel_file(
                    blob_content, document_name, container_name, blob_path, 
                    bot_id, user_id, file_extension
                )
            
            if not documents:
                raise ValueError(f"No processable data found in {file_extension} file")
            
            logger.debug(f"Successfully processed {file_extension} file: {len(documents)} documents created")
            return documents
            
        except Exception as e:
            logger.debug(f"Spreadsheet processing failed: {str(e)}")
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Spreadsheet processing failed: {str(e)}")

    async def _read_csv_file(self, blob_content, document_name):
        """Read CSV file with multiple encoding attempts"""
        # Try different encodings for CSV
        encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        for encoding in encodings:
            try:
                df = pd.read_csv(io.BytesIO(blob_content), encoding=encoding)
                logger.debug(f"Successfully read CSV with encoding: {encoding}")
                return df
            except (UnicodeDecodeError, pd.errors.EmptyDataError) as e:
                logger.debug(f"Failed to read CSV with encoding {encoding}: {str(e)}")
                continue
        
        raise ValueError("Could not decode CSV file with any supported encoding")

    async def _process_excel_file(self, blob_content, document_name, container_name, 
                                blob_path, bot_id, user_id, file_extension):
        """Process Excel file with multiple sheets, excluding hidden sheets"""
        documents = []
        
        try:
            # Read Excel file with all sheets
            excel_file = pd.ExcelFile(io.BytesIO(blob_content))
            
            # Get sheet information including visibility - THIS IS THE KEY FIX
            sheet_info = self._get_sheet_visibility_info(blob_content)
            
            logger.debug(f"Found {len(excel_file.sheet_names)} total sheets in {document_name}")
            logger.debug(f"Sheet visibility info: {sheet_info}")  # Debug logging
            
            # Create a workbook summary document
            workbook_summary = await self._create_workbook_summary(
                excel_file, sheet_info, document_name, container_name, 
                blob_path, bot_id, user_id, file_extension
            )
            documents.append(workbook_summary)
            
            # Process each visible sheet
            visible_sheet_count = 0
            for sheet_index, sheet_name in enumerate(excel_file.sheet_names):
                # Check if sheet is hidden - THIS IS THE CRITICAL CHECK
                sheet_visibility = sheet_info.get(sheet_name, {})
                is_hidden = sheet_visibility.get('hidden', False)
                
                if is_hidden:
                    logger.debug(f"Skipping hidden sheet: {sheet_name} (state: {sheet_visibility.get('sheet_state', 'unknown')})")
                    continue
                
                visible_sheet_count += 1
                logger.debug(f"Processing visible sheet {visible_sheet_count}: {sheet_name}")
                
                try:
                    # Read the specific sheet
                    df = pd.read_excel(io.BytesIO(blob_content), sheet_name=sheet_name)
                    
                    if df.empty:
                        logger.debug(f"Sheet '{sheet_name}' is empty, skipping...")
                        continue
                    
                    # Process this sheet
                    sheet_documents = await self._process_single_sheet(
                        df, document_name, sheet_name, container_name, blob_path,
                        bot_id, user_id, file_extension, sheet_index + 1, visible_sheet_count
                    )
                    documents.extend(sheet_documents)
                    
                except Exception as sheet_error:
                    logger.debug(f"Error processing sheet '{sheet_name}': {str(sheet_error)}")
                    # Create an error document for this sheet
                    error_doc = self._create_sheet_error_document(
                        sheet_name, str(sheet_error), container_name, blob_path,
                        bot_id, user_id, file_extension, sheet_index + 1
                    )
                    documents.append(error_doc)
                    continue
            
            logger.debug(f"Processed {visible_sheet_count} visible sheets from {document_name}")
            
        except Exception as e:
            logger.debug(f"Excel file processing failed: {str(e)}")
            raise ValueError(f"Could not read Excel file: {str(e)}")
        
        return documents

    def _get_sheet_visibility_info(self, blob_content):
        """Get information about sheet visibility using openpyxl"""
        sheet_info = {}
        
        try:
            # Use openpyxl to check sheet visibility
            from openpyxl import load_workbook
            
            # Create a fresh BytesIO object for openpyxl
            workbook_bytes = io.BytesIO(blob_content)
            
            # Load workbook with openpyxl to check sheet states
            workbook = load_workbook(workbook_bytes, read_only=True)
            
            logger.debug(f"Loaded workbook with {len(workbook.worksheets)} worksheets")
            
            for sheet in workbook.worksheets:
                sheet_state = getattr(sheet, 'sheet_state', 'visible')
                is_hidden = sheet_state != 'visible'
                
                sheet_info[sheet.title] = {
                    'hidden': is_hidden,
                    'sheet_state': sheet_state,
                    'index': workbook.worksheets.index(sheet)
                }
                
                logger.debug(f"Sheet '{sheet.title}': state='{sheet_state}', hidden={is_hidden}")
            
            workbook.close()
            
        except ImportError:
            logger.debug("Warning: openpyxl not available, cannot detect hidden sheets")
            # Fallback: assume all sheets are visible
            try:
                excel_file = pd.ExcelFile(io.BytesIO(blob_content))
                for i, sheet_name in enumerate(excel_file.sheet_names):
                    sheet_info[sheet_name] = {
                        'hidden': False,
                        'sheet_state': 'visible',
                        'index': i
                    }
                    logger.debug(f"Fallback: Sheet '{sheet_name}' assumed visible")
            except Exception as fallback_error:
                logger.debug(f"Fallback failed: {str(fallback_error)}")
        except Exception as e:
            logger.debug(f"Warning: Could not determine sheet visibility: {str(e)}")
            # Fallback: assume all sheets are visible
            try:
                excel_file = pd.ExcelFile(io.BytesIO(blob_content))
                for i, sheet_name in enumerate(excel_file.sheet_names):
                    sheet_info[sheet_name] = {
                        'hidden': False,
                        'sheet_state': 'visible',
                        'index': i
                    }
                    logger.debug(f"Error fallback: Sheet '{sheet_name}' assumed visible")
            except Exception as fallback_error:
                logger.debug(f"Error fallback failed: {str(fallback_error)}")
        
        return sheet_info

    async def _create_workbook_summary(self, excel_file, sheet_info, document_name, 
                                     container_name, blob_path, bot_id, user_id, file_extension):
        """Create a summary document for the entire Excel workbook"""
        
        visible_sheets = [name for name, info in sheet_info.items() if not info.get('hidden', False)]
        hidden_sheets = [name for name, info in sheet_info.items() if info.get('hidden', False)]
        
        summary_content = f"Excel Workbook: {document_name}\n"
        summary_content += f"File Type: {file_extension.upper()}\n"
        summary_content += f"Total Sheets: {len(excel_file.sheet_names)}\n"
        summary_content += f"Visible Sheets: {len(visible_sheets)}\n"
        summary_content += f"Hidden Sheets: {len(hidden_sheets)}\n\n"
        
        if visible_sheets:
            summary_content += "Visible Sheets:\n"
            for i, sheet_name in enumerate(visible_sheets, 1):
                summary_content += f"  {i}. {sheet_name}\n"
            summary_content += "\n"
        
        if hidden_sheets:
            summary_content += "Hidden Sheets (not processed):\n"
            for i, sheet_name in enumerate(hidden_sheets, 1):
                summary_content += f"  {i}. {sheet_name}\n"
            summary_content += "\n"
        
        # Add basic info about each visible sheet
        summary_content += "Sheet Details:\n"
        for sheet_name in visible_sheets:
            try:
                df = pd.read_excel(io.BytesIO(excel_file._io.getvalue()), sheet_name=sheet_name)
                summary_content += f"  {sheet_name}: {len(df)} rows, {len(df.columns)} columns\n"
                if len(df.columns) > 0:
                    summary_content += f"    Columns: {', '.join(df.columns.astype(str).tolist()[:5])}{'...' if len(df.columns) > 5 else ''}\n"
            except Exception as e:
                summary_content += f"  {sheet_name}: Error reading sheet - {str(e)}\n"
        
        # Extract keywords
        keywords = []
        try:
            from Managers.Document_Intelligence_3.semantic_chunker import extract_keywords
            keywords = extract_keywords(summary_content)
        except Exception:
            keywords = visible_sheets[:5]  # Use sheet names as fallback keywords
        
        metadata = {
            "source": f"{container_name}/{blob_path}",
            "chunk": 1,
            "page_number": 1,
            "extraction_time": datetime.datetime.now().isoformat(),
            "context": f"workbook summary - {file_extension}",
            "page_summary": f"Excel workbook with {len(visible_sheets)} visible sheets",
            "keywords": ", ".join(keywords) if keywords else "",
            "bot_id": bot_id if bot_id else f"default-bot-{container_name}",
            "file_type": file_extension,
            "total_sheets": len(excel_file.sheet_names),
            "visible_sheets": len(visible_sheets),
            "hidden_sheets": len(hidden_sheets),
            "sheet_names": visible_sheets
        }
        
        if user_id is not None:
            metadata["user_id"] = user_id
        
        return {
            "id": f"workbook_{uuid.uuid4().hex[:8]}",
            "content": summary_content,
            "metadata": metadata
        }

    async def _process_single_sheet(self, df, document_name, sheet_name, container_name, 
                                   blob_path, bot_id, user_id, file_extension, 
                                   sheet_index, visible_sheet_number):
        """Process a single sheet from Excel or CSV file"""
        documents = []
        
        # Clean column names (remove leading/trailing whitespace)
        df.columns = df.columns.astype(str).str.strip()
        
        # Create sheet summary
        summary_content = f"Sheet: {sheet_name}\n"
        summary_content += f"From Workbook: {document_name}\n"
        summary_content += f"Sheet Number: {visible_sheet_number}\n"
        summary_content += f"Total Rows: {len(df)}\n"
        summary_content += f"Total Columns: {len(df.columns)}\n"
        summary_content += f"Column Names: {', '.join(df.columns.tolist())}\n\n"
        
        # Add basic statistics for numeric columns
        numeric_columns = df.select_dtypes(include=['number']).columns.tolist()
        if numeric_columns:
            summary_content += f"Numeric Columns: {', '.join(numeric_columns)}\n"
            summary_content += "Basic Statistics:\n"
            try:
                stats = df[numeric_columns].describe()
                summary_content += stats.to_string() + "\n\n"
            except Exception:
                summary_content += "Could not generate statistics\n\n"
        
        # Add data preview (first few rows)
        preview_rows = min(5, len(df))
        summary_content += f"Data Preview (first {preview_rows} rows):\n"
        try:
            summary_content += df.head(preview_rows).to_string(index=False, max_cols=10)
        except Exception as e:
            logger.debug(f"Warning: Could not create data preview for sheet {sheet_name}: {str(e)}")
            summary_content += "Could not display data preview"
        
        # Extract keywords
        keywords = []
        try:
            from Managers.Document_Intelligence_3.semantic_chunker import extract_keywords
            keywords = extract_keywords(summary_content)
        except Exception:
            keywords = [sheet_name] + df.columns.tolist()[:5]
        
        # Create metadata for sheet summary
        metadata = {
            "source": f"{container_name}/{blob_path}",
            "chunk": visible_sheet_number + 1,  # +1 because workbook summary is chunk 1
            "page_number": visible_sheet_number + 1,
            "extraction_time": datetime.datetime.now().isoformat(),
            "context": f"sheet summary - {file_extension}",
            "page_summary": f"Sheet '{sheet_name}' with {len(df)} rows and {len(df.columns)} columns",
            "keywords": ", ".join(keywords) if keywords else "",
            "bot_id": bot_id if bot_id else f"default-bot-{container_name}",
            "file_type": file_extension,
            "sheet_name": sheet_name,
            "sheet_index": sheet_index,
            "visible_sheet_number": visible_sheet_number,
            "row_count": len(df),
            "column_count": len(df.columns),
            "columns": df.columns.tolist(),
            "numeric_columns": numeric_columns
        }
        
        if user_id is not None:
            metadata["user_id"] = user_id
        
        # Create sheet summary document
        sheet_summary_doc = {
            "id": f"sheet_{uuid.uuid4().hex[:8]}",
            "content": summary_content,
            "metadata": metadata
        }
        documents.append(sheet_summary_doc)
        
        # Create data chunks if dataset is large
        chunk_threshold = 50
        if len(df) > chunk_threshold:
            chunk_size = 50  # rows per chunk
            
            for i in range(0, len(df), chunk_size):
                chunk_df = df.iloc[i:i+chunk_size].copy()
                
                # Create content for this chunk
                chunk_content = f"Data chunk {i//chunk_size + 1} from sheet '{sheet_name}':\n"
                chunk_content += f"From workbook: {document_name}\n"
                chunk_content += f"Rows {i+1} to {min(i+chunk_size, len(df))} out of {len(df)} total rows:\n\n"
                
                try:
                    chunk_content += chunk_df.to_string(index=False, max_cols=10)
                except Exception as e:
                    logger.debug(f"Warning: Could not convert chunk to string: {str(e)}")
                    chunk_content += f"Data chunk contains {len(chunk_df)} rows with columns: {', '.join(chunk_df.columns.tolist())}"
                
                # Add summary statistics for this chunk
                chunk_numeric = chunk_df.select_dtypes(include=['number'])
                if not chunk_numeric.empty:
                    chunk_content += f"\n\nChunk Statistics:\n"
                    try:
                        chunk_content += chunk_numeric.describe().to_string()
                    except Exception:
                        chunk_content += "Could not generate statistics for this chunk"
                
                # Extract keywords for this chunk
                chunk_keywords = []
                try:
                    from Managers.Document_Intelligence_3.semantic_chunker import extract_keywords
                    chunk_keywords = extract_keywords(chunk_content)
                except Exception:
                    chunk_keywords = [sheet_name] + chunk_df.columns.tolist()[:3]
                
                # Create metadata for this chunk
                chunk_metadata = {
                    "source": f"{container_name}/{blob_path}",
                    "chunk": len(documents) + visible_sheet_number + 1,
                    "page_number": len(documents) + visible_sheet_number + 1,
                    "extraction_time": datetime.datetime.now().isoformat(),
                    "context": f"sheet data chunk - {file_extension}",
                    "page_summary": f"Sheet '{sheet_name}' rows {i+1} to {min(i+chunk_size, len(df))} of {len(df)} total",
                    "keywords": ", ".join(chunk_keywords) if chunk_keywords else "",
                    "bot_id": bot_id if bot_id else f"default-bot-{container_name}",
                    "file_type": file_extension,
                    "sheet_name": sheet_name,
                    "sheet_index": sheet_index,
                    "visible_sheet_number": visible_sheet_number,
                    "chunk_start_row": i + 1,
                    "chunk_end_row": min(i + chunk_size, len(df)),
                    "columns": chunk_df.columns.tolist(),
                    "chunk_row_count": len(chunk_df)
                }
                
                if user_id is not None:
                    chunk_metadata["user_id"] = user_id
                
                chunk_doc = {
                    "id": f"chunk_{uuid.uuid4().hex[:8]}",
                    "content": chunk_content,
                    "metadata": chunk_metadata
                }
                documents.append(chunk_doc)
        
        return documents

    def _create_sheet_error_document(self, sheet_name, error_message, container_name, 
                                    blob_path, bot_id, user_id, file_extension, sheet_index):
        """Create an error document for sheets that couldn't be processed"""
        
        error_content = f"Error processing sheet: {sheet_name}\n"
        error_content += f"Error message: {error_message}\n"
        error_content += f"Sheet index: {sheet_index}\n"
        error_content += "This sheet could not be processed due to the above error."
        
        metadata = {
            "source": f"{container_name}/{blob_path}",
            "chunk": sheet_index + 1,
            "page_number": sheet_index + 1,
            "extraction_time": datetime.datetime.now().isoformat(),
            "context": f"sheet error - {file_extension}",
            "page_summary": f"Error processing sheet '{sheet_name}'",
            "keywords": f"error, {sheet_name}",
            "bot_id": bot_id if bot_id else f"default-bot-{container_name}",
            "file_type": file_extension,
            "sheet_name": sheet_name,
            "sheet_index": sheet_index,
            "error": True,
            "error_message": error_message
        }
        
        if user_id is not None:
            metadata["user_id"] = user_id
        
        return {
            "id": f"error_{uuid.uuid4().hex[:8]}",
            "content": error_content,
            "metadata": metadata
        }

    async def _process_with_document_intelligence(self, blob_content, document_name, container_name, 
                                                blob_path, model, headers_to_split_on, bot_id, user_id):
        """Process documents using Azure Document Intelligence (original logic)"""
        temp_file_path = None
        
        try:
            # Create a temporary file
            with tempfile.NamedTemporaryFile(delete=False) as temp_file:
                temp_file.write(blob_content)
                temp_file_path = temp_file.name
            
            # Use the Document Intelligence client directly for better page separation
            with open(temp_file_path, "rb") as f:
                file_content = f.read()
                poller = self.client.begin_analyze_document(
                    model_id=model,
                    body=file_content,
                    content_type="application/octet-stream"
                )
                result = poller.result()
            
            # Process the document page by page
            documents = []
            for page_index, page in enumerate(result.pages):
                # Extract text content for this page
                if not page.lines:
                    logger.debug(f"Warning: No lines found on page {page.page_number}")
                    continue  # Skip processing this page
                    
                page_lines = [line.content for line in page.lines]
                page_content = '\n'.join(page_lines)
                
                # Generate a unique ID for this page
                page_id = f"doc_{uuid.uuid4().hex[:8]}"
                
                # Try to extract keywords
                keywords = []
                try:
                    from Managers.Document_Intelligence_3.semantic_chunker import extract_keywords
                    keywords = extract_keywords(page_content)
                except Exception as e:
                    logger.debug(f"Warning: Could not extract keywords: {str(e)}")
                
                # Create metadata with all the common fields
                metadata = {
                    "source": f"{container_name}/{blob_path}",
                    "chunk": page_index + 1,
                    "page_number": page.page_number,
                    "extraction_time": datetime.datetime.now().isoformat(),
                    "context": "document page",
                    "page_summary": "",
                    "keywords": ", ".join(keywords) if keywords else "",
                    "bot_id": bot_id if bot_id else f"default-bot-{container_name}",
                    "file_type": document_name.lower().split('.')[-1] if '.' in document_name else "unknown"
                }
                
                # Add user_id to metadata if provided
                if user_id is not None:
                    metadata["user_id"] = user_id
                
                # Create a document for this page with all the fields
                page_doc = {
                    "id": page_id,
                    "content": page_content,
                    "metadata": metadata
                }
                
                documents.append(page_doc)
            
            # If we need to split by headers
            if headers_to_split_on:
                all_splits = []
                text_splitter = MarkdownHeaderTextSplitter(headers_to_split_on=headers_to_split_on)
                
                for doc_index, doc in enumerate(documents):
                    try:
                        # Split the content by headers
                        splits = text_splitter.split_text(doc["content"])
                        
                        # Add each split with the correct page metadata
                        for split_index, split in enumerate(splits):
                            # Try to extract keywords for this split
                            split_keywords = []
                            try:
                                from Managers.Document_Intelligence_3.semantic_chunker import extract_keywords
                                split_keywords = extract_keywords(split.page_content)
                            except Exception:
                                pass
                            
                            # Start with base metadata
                            combined_metadata = {
                                "source": doc["metadata"]["source"],
                                "chunk": split_index + 1,
                                "page_number": doc["metadata"]["page_number"],
                                "extraction_time": doc["metadata"]["extraction_time"],
                                "context": "header section",
                                "page_summary": "",
                                "keywords": ", ".join(split_keywords) if split_keywords else "",
                                "bot_id": doc["metadata"]["bot_id"],
                                "file_type": doc["metadata"]["file_type"]
                            }
                            
                            # Add user_id to metadata if it exists in the original metadata
                            if "user_id" in doc["metadata"]:
                                combined_metadata["user_id"] = doc["metadata"]["user_id"]
                            
                            # Generate a unique ID for this chunk
                            chunk_id = f"chunk_{uuid.uuid4().hex[:8]}"
                            
                            all_splits.append({
                                "id": chunk_id,
                                "content": split.page_content,
                                "metadata": combined_metadata
                            })
                    except Exception as split_error:
                        logger.debug(f"Error splitting page {doc['metadata']['page_number']}: {str(split_error)}")
                        # If splitting fails, use the whole page
                        all_splits.append(doc)
                
                return all_splits
            
            # If not splitting by headers, return the page-separated documents
            return documents
            
        except Exception as e:
            logger.debug(f"Document Intelligence processing failed: {str(e)}")
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Document Intelligence processing failed: {str(e)}")
            
        finally:
            # Clean up temp file
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.unlink(temp_file_path)
                except Exception as e:
                    logger.debug(f"Warning: Could not delete temp file {temp_file_path}: {str(e)}")

    @profile_endpoint(name='analyze_document_from_file')
    async def analyze_document_from_file(self, file_content: bytes, file_name: str, model="prebuilt-layout"):
        """
        Analyze a document from file content using Document Intelligence

        Args:
            file_content: Raw file content as bytes
            file_name: Name of the file
            model: Document intelligence model to use

        Returns:
            Analysis result from Document Intelligence
        """
        temp_file_path = None
        file_extension = file_name.lower().split('.')[-1] if '.' in file_name else 'unknown'
        doc_size_kb = len(file_content) / 1024 if file_content else 0

        # Update decorator metadata - this is a workaround since decorators are applied first
        # The context manager inside will provide more accurate profiling

        try:
            # Determine appropriate file extension for temp file
            file_extension = file_name.lower().split('.')[-1] if '.' in file_name else ''
            suffix = f'.{file_extension}' if file_extension else ''
            
            # Create a temporary file with appropriate extension
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp_file:
                temp_file.write(file_content)
                temp_file_path = temp_file.name
            
            # Analyze the document
            with open(temp_file_path, "rb") as f:
                poller = self.client.begin_analyze_document(
                    model_id=model,
                    body=f.read(),
                    content_type="application/octet-stream"
                )
                result = poller.result()
            
            return result
            
        except Exception as e:
            logger.debug(f"Document analysis failed for {file_name}: {str(e)}")
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Document analysis failed: {str(e)}")
            
        finally:
            # Clean up temp file
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.unlink(temp_file_path)
                except Exception as e:
                    logger.debug(f"Warning: Could not delete temp file {temp_file_path}: {str(e)}")

    @profile_endpoint(name='extract_word_document_metadata', document_type='docx')
    async def extract_word_document_metadata(self, file_content: bytes, file_name: str) -> Dict[str, Any]:
        """
        Extract comprehensive metadata from Word documents

        Args:
            file_content: Raw file content as bytes
            file_name: Name of the file

        Returns:
            Dictionary containing document metadata
        """
        try:
            result = await self.analyze_document_from_file(file_content, file_name)
            
            metadata = {
                "file_name": file_name,
                "file_type": "Microsoft Word Document",
                "total_pages": len(result.pages) if result.pages else 0,
                "total_paragraphs": len(result.paragraphs) if result.paragraphs else 0,
                "has_tables": bool(result.tables),
                "table_count": len(result.tables) if result.tables else 0,
                "extraction_time": datetime.datetime.now().isoformat(),
                "document_structure": {}
            }
            
            # Analyze document structure
            if result.paragraphs:
                styles = set()
                headers = []
                content_length = 0
                
                for para in result.paragraphs:
                    if hasattr(para, 'role') and para.role:
                        styles.add(para.role)
                        if para.role in ['title', 'sectionHeading']:
                            headers.append(para.content.strip())
                    
                    content_length += len(para.content)
                
                metadata["document_structure"] = {
                    "styles_found": list(styles),
                    "headers": headers[:10],  # First 10 headers
                    "total_content_length": content_length,
                    "average_paragraph_length": content_length / len(result.paragraphs) if result.paragraphs else 0
                }
            
            # Table information
            if result.tables:
                table_info = []
                for i, table in enumerate(result.tables):
                    table_metadata = {
                        "table_index": i,
                        "row_count": getattr(table, 'row_count', 'unknown'),
                        "column_count": getattr(table, 'column_count', 'unknown'),
                        "cell_count": len(table.cells) if table.cells else 0
                    }
                    table_info.append(table_metadata)
                
                metadata["tables"] = table_info
            
            return metadata
            
        except Exception as e:
            logger.debug(f"Metadata extraction failed for {file_name}: {str(e)}")
            return {
                "file_name": file_name,
                "file_type": "Microsoft Word Document",
                "error": str(e),
                "extraction_time": datetime.datetime.now().isoformat()
            }

    @track_document_processing("DocumentIntelligence")
    @profile_endpoint(name='process_spreadsheet_file', document_type='spreadsheet')
    async def process_spreadsheet_file(self, container_name: str, blob_path: str, bot_id: str,
                                     user_id: Optional[str] = None, headers_to_split_on: Optional[bool] = False):
        """
        Process CSV and Excel files from blob storage

        Args:
            container_name: Azure Storage container name
            blob_path: Path to the blob within the container
            bot_id: ID of the bot for multi-tenancy filtering
            user_id: ID of the user for multi-tenancy filtering
            headers_to_split_on: Whether to split by headers (for CSV files)

        Returns:
            List of processed document chunks ready for embedding
        """
        try:
            import logging
            logger = logging.getLogger(__name__)
            
            logger.info(f"DEBUG: Processing spreadsheet file - container_name: {container_name}, blob_path: {blob_path}, bot_id: {bot_id}, user_id: {user_id}, headers_to_split_on: {headers_to_split_on}")
            
            # Debug storage manager
            logger.info(f"DEBUG: storage_manager type: {type(self.storage_manager)}")
            logger.info(f"DEBUG: storage_manager has download_file: {hasattr(self.storage_manager, 'download_file')}")
            logger.info(f"DEBUG: storage_manager has download_blob: {hasattr(self.storage_manager, 'download_blob')}")
            
            # Get file content from blob storage
            blob_content = await self.storage_manager.download_file(container_name, blob_path)
            logger.info(f"DEBUG: Successfully downloaded blob content, size: {len(blob_content) if blob_content else 0} bytes")
            
            # Get file extension
            file_extension = blob_path.lower().split('.')[-1] if '.' in blob_path else ''
            
            # Process the spreadsheet file
            documents = await self._process_spreadsheet_file(
                blob_content, 
                blob_path.split('/')[-1],  # Get filename
                file_extension, 
                container_name, 
                blob_path, 
                bot_id, 
                user_id, 
                headers_to_split_on
            )
            
            return documents
            
        except Exception as e:
            logger.debug(f"Error processing spreadsheet file: {str(e)}")
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Error processing spreadsheet file: {str(e)}")


# Dependency function to get DocumentIntelligence instance
def get_document_intelligence_manager(settings: Settings = Depends(get_settings)) -> DocumentIntelligence:
    """Dependency function to get DocumentIntelligence instance"""
    return DocumentIntelligence(settings)